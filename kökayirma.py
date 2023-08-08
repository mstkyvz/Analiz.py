from jpype import JClass, getDefaultJVMPath, shutdownJVM, startJVM
from openpyxl import load_workbook
from pathlib import Path

# Zemberek JAR dosyasının yolu
ZEMBEREK_PATH = r"C:\Users\Mustafa\Desktop\zemberek-full.jar"

# Zemberek sınıflarını yükleyin
startJVM(getDefaultJVMPath(), "-ea", "-Djava.class.path=" + ZEMBEREK_PATH)

# Zemberek Türkiye Türkçesi için gerekli sınıfları alın
TurkishMorphology = JClass("zemberek.morphology.TurkishMorphology")
morphology = TurkishMorphology.createWithDefaults()

# Kök analizini gerçekleştiren fonksiyon
def analyze_word(word):
    results = morphology.analyzeAndDisambiguate(word).bestAnalysis()
    if results:
        return results[0]

# Excel dosyasını aç
file_path = Path("Kelimeler.xlsx")
workbook = load_workbook(file_path)
sheet = workbook.active

# Her hücreyi dolaş ve kök analizi yaparak sonuçları yaz
for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    if row[0]:
        word = row[0]
        analysis = analyze_word(word)
        if analysis:
            lemma = analysis.getLemmas()[0]
            sheet.cell(row=idx, column=2, value=str(lemma))  # Zemberek çıktısını metne dönüştür
            # Ekleri ve kökü ayrı sütunlara ayır
            stem = str(analysis.getStem())
            sheet.cell(row=idx, column=2, value=str(stem))  # Kök sütunu
            sheet.cell(row=idx, column=3, value=str(row[0][len(stem):]))  # Ek sütunu

# Sonuçları kaydet
output_path = Path("Kelime_kök_ekler_ayir.xlsx")
workbook.save(output_path)

# JVM'yi kapat
shutdownJVM()
